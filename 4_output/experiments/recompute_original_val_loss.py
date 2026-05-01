import argparse
import json
import os
import sys
import time
import warnings
from pathlib import Path

import numpy as np
import optuna
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import PowerTransformer


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "2_preprocessing"))
sys.path.insert(0, str(ROOT / "3_modeling"))

from utils.config import TARGET_COL  # noqa: E402
from utils.data import get_feat_cols, load_all, split_xs  # noqa: E402
from utils.evaluate import rmse  # noqa: E402
from modules.e2e_hpo import rerun_best_trial_with_pp  # noqa: E402


def find_summary_file() -> Path:
    matches = sorted(ROOT.rglob("experiment_status_summary.xlsx"))
    if not matches:
        raise FileNotFoundError("experiment_status_summary.xlsx not found")
    return matches[0]


def sqlite_storage(db_path: Path) -> str:
    return "sqlite:///" + db_path.resolve().as_posix()


def load_summary_rows(summary_path: Path) -> list[dict]:
    wb = load_workbook(summary_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row.get("exp_id"):
            rows.append(row)
    return rows


def build_target_space(ys: dict, transform: str, clip_y_extreme: bool):
    ys_input = {k: v.copy() for k, v in ys.items()}

    if clip_y_extreme:
        train_y = ys_input["train"][TARGET_COL]
        second_max = train_y[train_y < train_y.max()].max()
        ys_input["train"][TARGET_COL] = train_y.clip(upper=second_max)

    if transform == "log1p":
        for df in ys_input.values():
            if TARGET_COL in df.columns:
                df[TARGET_COL] = np.log1p(df[TARGET_COL].values)

        def inverse(arr):
            return np.clip(np.expm1(np.asarray(arr, dtype=float)), 0.0, None)

        return ys_input, inverse

    if transform == "yeo-johnson":
        transformer = PowerTransformer(method="yeo-johnson", standardize=False)
        y_fit = ys_input["train"][TARGET_COL].values.reshape(-1, 1)
        transformer.fit(y_fit)
        z_fit = transformer.transform(y_fit).ravel()
        z_min = float(np.nanmin(z_fit))
        z_max = float(np.nanmax(z_fit))

        for df in ys_input.values():
            if TARGET_COL in df.columns:
                df[TARGET_COL] = transformer.transform(
                    df[TARGET_COL].values.reshape(-1, 1)
                ).ravel()

        def inverse(arr):
            arr = np.clip(np.asarray(arr, dtype=float), z_min, z_max)
            out = transformer.inverse_transform(arr.reshape(-1, 1)).ravel()
            out = np.clip(out, 0.0, None)
            return np.nan_to_num(out, nan=0.0, posinf=0.0, neginf=0.0)

        return ys_input, inverse

    def identity(arr):
        return np.asarray(arr, dtype=float)

    return ys_input, identity


def recompute_one(exp_id: str, study, xs, xs_dict, ys, feat_cols) -> dict:
    trial = study.best_trial
    ua = study.user_attrs
    transform = ua.get("target_transform", "none")
    clip_y_extreme = bool(ua.get("clip_y_extreme", False))
    ys_input, inverse = build_target_space(ys, transform, clip_y_extreme)
    e2e = ua["e2e_params"]

    start = time.time()
    final = rerun_best_trial_with_pp(
        xs=xs,
        xs_dict=xs_dict,
        ys=ys_input,
        feat_cols=feat_cols,
        best_params=trial.params,
        best_pp_params_resolved=trial.user_attrs.get("resolved_pp_params"),
        pipeline_config=ua["pipeline_config"],
        clf_model=e2e["clf_model"],
        reg_model=e2e["reg_model"],
        label_col=ua.get("label_col", "label_bin"),
        imbalance_method=e2e["imbalance_method"],
        top_k_fixed=e2e["top_k_fixed"],
        clf_filter_threshold_fixed=e2e["clf_filter_threshold_fixed"],
        zero_clip_threshold_fixed=e2e["zero_clip_threshold_fixed"],
        clf_fixed=e2e.get("clf_fixed"),
        reg_fixed=e2e.get("reg_fixed"),
        use_sampling=ua["sampling_params"].get("use_sampling", False),
        sample_frac=ua["sampling_params"].get("sample_frac", 1.0),
        exclude_cols=ua.get("exclude_cols", []) or [],
        **dict(ua["rerun_params"]),
    )

    y_val_raw = ys["validation"][TARGET_COL].values
    val_pred_raw = inverse(final["val_pred"])
    val_rmse_original = float(rmse(y_val_raw, val_pred_raw))

    result = {
        "exp_id": exp_id,
        "target_transform": transform,
        "clip_y_extreme": clip_y_extreme,
        "best_trial": trial.number,
        "hpo_objective_transformed": float(trial.value),
        "hpo_val_rmse_transformed": trial.user_attrs.get("val_rmse"),
        "rerun_val_rmse_transformed": final.get("val_rmse"),
        "val_rmse_original_recomputed": val_rmse_original,
        "elapsed_sec": round(time.time() - start, 2),
    }
    return result


def write_excel(summary_rows: list[dict], recomputed: pd.DataFrame, out_path: Path):
    summary = pd.DataFrame(summary_rows)
    merged = summary.merge(recomputed, on="exp_id", how="left")
    if "val_rmse" in merged.columns:
        merged = merged.rename(columns={"val_rmse": "summary_val_rmse_transformed"})
    merged["val_rmse"] = merged["val_rmse_original_recomputed"]
    merged = merged.sort_values(
        ["val_rmse", "summary_val_rmse_transformed"],
        na_position="last",
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="summary_recomputed", index=False)
        recomputed.sort_values("val_rmse_original_recomputed").to_excel(
            writer, sheet_name="recomputed_only", index=False
        )
        pd.DataFrame(
            [
                {
                    "key": "val_rmse",
                    "value": "original-scale validation RMSE recomputed from DB params and raw data",
                },
                {
                    "key": "summary_val_rmse_transformed",
                    "value": "original spreadsheet value; usually transformed-space HPO objective fallback",
                },
                {
                    "key": "checkpoint",
                    "value": str(out_path.with_suffix(".csv")),
                },
            ]
        ).to_excel(writer, sheet_name="notes", index=False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--only", nargs="*", default=None, help="exp_id list to recompute")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--force", action="store_true", help="recompute existing checkpoint rows")
    args = parser.parse_args()

    warnings.filterwarnings("ignore")
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")

    summary_path = find_summary_file()
    db_path = summary_path.parent / "optuna_merged.db"
    checkpoint = summary_path.parent / "experiment_status_summary_original_loss_recomputed.csv"
    out_xlsx = summary_path.parent / "experiment_status_summary_original_loss_recomputed.xlsx"

    rows = load_summary_rows(summary_path)
    if args.only:
        only = set(args.only)
        rows = [r for r in rows if str(r["exp_id"]) in only]
    if args.limit:
        rows = rows[: args.limit]

    done = pd.DataFrame()
    if checkpoint.exists() and not args.force:
        done = pd.read_csv(checkpoint)
        done_ids = set(done["exp_id"].astype(str))
    else:
        done_ids = set()

    print(f"summary={summary_path}")
    print(f"db={db_path}")
    print(f"checkpoint={checkpoint}")
    print(f"target studies={len(rows)}, already done={len(done_ids)}")

    print("[load] raw data")
    xs, ys = load_all()
    xs_dict = split_xs(xs)
    feat_cols = get_feat_cols(xs)

    results = done.to_dict("records") if not done.empty else []
    storage = sqlite_storage(db_path)

    for idx, row in enumerate(rows, start=1):
        exp_id = str(row["exp_id"])
        if exp_id in done_ids:
            print(f"[skip] {idx}/{len(rows)} {exp_id}")
            continue
        print(f"[run] {idx}/{len(rows)} {exp_id}")
        study = optuna.load_study(study_name=exp_id, storage=storage)
        try:
            result = recompute_one(exp_id, study, xs, xs_dict, ys, feat_cols)
            results.append(result)
            pd.DataFrame(results).to_csv(checkpoint, index=False, encoding="utf-8-sig")
            write_excel(rows, pd.DataFrame(results), out_xlsx)
            print(
                f"[ok] {exp_id} original={result['val_rmse_original_recomputed']:.9f} "
                f"elapsed={result['elapsed_sec']}s"
            )
        except Exception as exc:
            result = {
                "exp_id": exp_id,
                "error": repr(exc),
            }
            results.append(result)
            pd.DataFrame(results).to_csv(checkpoint, index=False, encoding="utf-8-sig")
            write_excel(rows, pd.DataFrame(results), out_xlsx)
            print(f"[error] {exp_id}: {exc!r}")

    if results:
        write_excel(rows, pd.DataFrame(results), out_xlsx)
        print(f"[saved] {out_xlsx}")


if __name__ == "__main__":
    main()
